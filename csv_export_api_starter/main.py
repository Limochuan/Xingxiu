#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os, io, re, datetime as dt
from typing import List, Dict, Any, Optional, Tuple
from urllib.parse import quote, urlencode, unquote

import pymysql
from fastapi import FastAPI, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse, RedirectResponse, HTMLResponse
from uvicorn import run as uvicorn_run
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 配置
# =========================
DB_HOST = os.getenv("DB_HOST", "rm-k1a5w7qk9cnm74r25wo.mysql.ap-southeast-5.rds.aliyuncs.com")
DB_PORT = int(os.getenv("DB_PORT", "3306"))
DB_USER = os.getenv("DB_USER", "script_xingxiu")
DB_PASS = os.getenv("DB_PASS", "Julong678678678")
DB_NAME = os.getenv("DB_NAME", "xingxiu_db")
DB_TABLE = os.getenv("DB_TABLE", "xingxiu_daily_report")

DEFAULT_PROJECT = "Kalteng GIJ 中加一园"
FONT_TITLE_NAME = "仿宋"
FONT_BODY_NAME  = "仿宋"

# 错误跳转域名（可用环境变量覆盖）
REDIRECT_BEFORE_CUTOFF = os.getenv("REDIRECT_BEFORE_CUTOFF", "http://www.info.julongairoxy.com/")
REDIRECT_TODAY_FUTURE  = os.getenv("REDIRECT_TODAY_FUTURE",  "http://www.info2.julongairoxy.com/")
REDIRECT_OTHER_ERROR   = os.getenv("REDIRECT_OTHER_ERROR",   "http://www.info1.julongairoxy.com/")
CUTOFF_DATE_STR = os.getenv("CUTOFF_DATE", "2025-09-01")
CUTOFF_DATE = dt.datetime.strptime(CUTOFF_DATE_STR, "%Y-%m-%d").date()

# 支持的项目名白名单（用于智能匹配）
PROJECT_WHITELIST = [
    "Kalteng GIJ 中加一园",
    "Kasel PBB 南加二园",
    "Kalsel PU 南加四园",
    "Kalteng ASP 中加七园",
    "Kalteng KLM 中加十一园",
]

app = FastAPI(title="Yaoguang Excel Download API", version="2.1.0")

# =========================
# 基础列清单（去掉 ID，保留全量供子表过滤/重排）
# =========================
# 结构： (英文字段名, 中文名, 备选字段名列表)
COLUMN_SPECS = [
    ("RANK_TODAY","当日排名",["RANK_TODAY"]),                 # 虚拟列，运行时写入
    ("DEVICE_ID","工时通编号",["DEVICE_ID"]),
    ("PROJECT_NAME","系统项目名称",["PROJECT_NAME"]),
    ("MECHANICAL_NO","系统编号",["MECHANICAL_NO"]),
    ("CAR_TYPE","系统类型",["CAR_TYPE"]),
    ("DATE_STR","数据日期",["DATE_STR"]),
    ("RENT_TYPE","系统计算类型",["RENT_TYPE"]),
    ("VALID_DURATION","有效工时",["VALID_DURATION"]),
    ("IDLING_DURATION","怠速工时",["IDLING_DURATION"]),
    ("VALID_PERCENT","工时有效比%",["VALID_PERCENT"]),
    ("DAY_OIL","油耗",["DAY_OIL"]),
    ("DAY_REFUEL","加油",["DAY_REFUEL"]),
    ("DAY_MILEAGE","里程",["DAY_MILEAGE"]),
    ("WORKHOUR_AVG_OIL","工时平均油耗",["WORKHOUR_AVG_OIL"]),
    ("TRANSPORT_AVG_OIL","运输平均油耗",["TRANSPORT_AVG_OIL"]),
    ("COMPANY","公司",["COMPANY"]),
    ("ESTATE","区域",["ESTATE"]),
    ("MACHINE_TYPE","机械类型",["MACHINE_TYPE"]),
    ("MACHINE_CATEGORY","机械具体类型",["MACHINE_CATEGORY"]),
    ("MACHINE_NO","园区编号",["MACHINE_NO"]),
    ("BRAND_SPEC","具体型号",["BRAND_SPEC"]),
    ("ORDER_NO SAP","SAP订单号",["ORDER_NO SAP","ORDER_NO_SAP","ORDER_NO"]),
    ("PURCHASE_DATE","购买日期",["PURCHASE_DATE"]),
    ("DRIVER_COUNT","司机数量",["DRIVER_COUNT"]),
    ("PURCH_PRICE","购买价格",["PURCH_PRICE","PURCHASE_PRICE"]),
    ("FUEL_DIFF","标准油耗差",["FUEL_DIFF"]),
    ("INSERT_TIME","数据插入时间",["INSERT_TIME","CREATE_TIME"]),
    ("SCORE","得分",["SCORE"]),
    ("SUMMARY","AI分析",["SUMMARY"]),
    ("ANALYSIS_TIME","分析时间",["ANALYSIS_TIME"]),
    ("MODEL_NAME","使用模型",["MODEL_NAME"]),
]

# =========================
# 工具函数
# =========================
def connect_db():
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
    )

def parse_date_like(s: Optional[str]) -> Optional[dt.date]:
    if not s: return None
    for fmt in ("%Y-%m-%d","%Y.%m.%d","%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s.strip(), fmt).date()
        except:
            pass
    m = re.match(r"^\s*(\d{4})[-./](\d{1,2})[-./](\d{1,2})\s*$", s or "")
    if m:
        y,mo,d = map(int, m.groups())
        return dt.date(y,mo,d)
    return None

def parse_date_range(raw: Optional[str]) -> Tuple[Optional[dt.date], Optional[dt.date]]:
    if not raw: return None, None
    s = raw.strip().replace("—","-").replace("–","-").replace("到","-").replace("to","-").replace("TO","-")
    s = re.sub(r"-{2,}", "-", s)
    parts = [p for p in s.split("-") if p.strip()]
    if len(parts) >= 2:
        return parse_date_like(parts[0]), parse_date_like(parts[-1])
    return None, None

def span_from_params(DATE_STR: Optional[str], DATE_FROM: Optional[str], DATE_TO: Optional[str], DATE_RANGE: Optional[str]) -> Tuple[Optional[dt.date], Optional[dt.date]]:
    if DATE_STR:
        d = parse_date_like(DATE_STR)
        return (d, d) if d else (None, None)
    left = parse_date_like(DATE_FROM) if DATE_FROM else None
    right = parse_date_like(DATE_TO) if DATE_TO else None
    if not (left and right) and DATE_RANGE:
        l2, r2 = parse_date_range(DATE_RANGE)
        left, right = left or l2, right or r2
    if left and right:
        if left > right: left, right = right, left
        return left, right
    return None, None

def choose_redirect_by_span(span: Tuple[Optional[dt.date], Optional[dt.date]]) -> str:
    today = dt.date.today()
    s, e = span
    if (s and s < CUTOFF_DATE) or (e and e < CUTOFF_DATE):
        return REDIRECT_BEFORE_CUTOFF
    if (s and s >= today) or (e and e >= today):
        return REDIRECT_TODAY_FUTURE
    return REDIRECT_OTHER_ERROR

def want_html(request: Request) -> bool:
    accept = (request.headers.get("accept") or "").lower()
    ua = (request.headers.get("user-agent") or "").lower()
    return "text/html" in accept or "mozilla" in ua

def redirect_with_error(msg: str, extra: dict, span: Tuple[Optional[dt.date], Optional[dt.date]], status_code: int = 303):
    params = {"err": msg}
    for k in ["DATE_STR","DATE_FROM","DATE_TO","DATE_RANGE","PROJECT_NAME"]:
        v = extra.get(k)
        if v: params[k] = v
    base = choose_redirect_by_span(span)
    return RedirectResponse(url=f"{base}?{urlencode(params, safe='')}", status_code=status_code)

def _normalize_name(s: str) -> str:
    if s is None: return ""
    s2 = re.sub(r"\s+", "", s)
    return s2.casefold()

def normalize_project_name(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    decoded = unquote(raw).strip()
    if not decoded:
        return None
    canon_map = { _normalize_name(p): p for p in PROJECT_WHITELIST }
    hit = canon_map.get(_normalize_name(decoded))
    return hit or decoded

def detect_date_span_from_rows(rows: List[Dict[str, Any]]) -> Tuple[Optional[dt.date], Optional[dt.date]]:
    dates = []
    for r in rows:
        v = r.get("DATE_STR")
        if isinstance(v, dt.datetime): dates.append(v.date())
        elif isinstance(v, dt.date):   dates.append(v)
        elif isinstance(v, str):
            d = parse_date_like(v)
            if d: dates.append(d)
    return (min(dates), max(dates)) if dates else (None, None)

def _first_present(row: Dict[str, Any], alts: list) -> Any:
    for k in alts:
        if k in row and row[k] is not None:
            return row[k]
    return ""

def _as_float(x, default: float) -> float:
    try:
        if x is None: return default
        if isinstance(x, (int, float)): return float(x)
        s = str(x).strip()
        if not s: return default
        return float(s)
    except:
        return default

# =========================
# 分类 & 排名
# =========================
def classify_alat(row: Dict[str, Any]) -> Optional[str]:
    keys = ("MACHINE_CATEGORY", "MACHINE_TYPE", "CAR_TYPE")
    text = " ".join([str(row.get(k, "")) for k in keys]).casefold()
    if "sedang" in text:
        return "sedang"
    if "berat" in text:
        return "berat"
    return None

def sort_key_for_ranking(row: Dict[str, Any]):
    score = _as_float(row.get("SCORE"), float("-inf"))
    valid = _as_float(row.get("VALID_DURATION"), float("-inf"))
    oil   = _as_float(row.get("DAY_OIL"), float("-inf"))
    try:
        rid = int(row.get("ID")) if row.get("ID") is not None else 10**12
    except:
        rid = 10**12
    return (-score, -valid, -oil, rid)

def sort_and_rank_by_date(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # 按 DATE_STR 分桶，每桶内排序并写入 RANK_TODAY = 1..n
    buckets: Dict[str, List[Dict[str, Any]]] = {}
    def norm_date_str(v) -> str:
        if isinstance(v, dt.datetime): return v.date().strftime("%Y-%m-%d")
        if isinstance(v, dt.date):     return v.strftime("%Y-%m-%d")
        if isinstance(v, str):
            d = parse_date_like(v)
            return d.strftime("%Y-%m-%d") if d else "0001-01-01"
        return "0001-01-01"

    for r in rows:
        buckets.setdefault(norm_date_str(r.get("DATE_STR")), []).append(r)

    out: List[Dict[str, Any]] = []
    for date_key in sorted(buckets.keys()):
        group = buckets[date_key]
        group_sorted = sorted(group, key=sort_key_for_ranking)
        for idx, row in enumerate(group_sorted, start=1):
            row = dict(row)
            row["RANK_TODAY"] = idx
            out.append(row)
    return out

# =========================
# 列过滤/重排（按子表）
# =========================
def build_columns_for_sheet(sheet_kind: str) -> List[Tuple[str,str,list]]:
    """
    sheet_kind: 'sedang' or 'berat'
    规则：
      - 先放固定头：RANK_TODAY, MACHINE_NO, SCORE, SUMMARY
      - 再接“其余未隐藏字段”的原相对顺序
      - 隐藏字段按需求表
    """
    hidden_sedang = {
        "MECHANICAL_NO", "VALID_DURATION", "IDLING_DURATION",
        "VALID_PERCENT", "WORKHOUR_AVG_OIL",
    }
    hidden_berat = {
        "MECHANICAL_NO", "DAY_MILEAGE", "TRANSPORT_AVG_OIL",
    }

    hidden = hidden_sedang if sheet_kind == "sedang" else hidden_berat
    head_order = ["RANK_TODAY", "MACHINE_NO", "SCORE", "SUMMARY"]

    # 映射方便查
    spec_map = {en: spec for en, zh, alts in COLUMN_SPECS for spec in [(en, zh, alts)]}
    # 先挑头部
    chosen = []
    used = set()
    for key in head_order:
        if key in hidden: 
            continue
        en, zh, alts = spec_map[key]
        chosen.append((en, zh, alts))
        used.add(key)

    # 再补其余（排除隐藏 & 已使用）
    for en, zh, alts in COLUMN_SPECS:
        if en in hidden or en in used:
            continue
        chosen.append((en, zh, alts))
        used.add(en)

    return chosen

def col_widths_for_specs(ncols: int) -> List[float]:
    # 给第一列排名稍宽，其他用一个合理通用宽度（保守）
    widths = [10] + [18]*(ncols-1)
    return widths

# =========================
# Excel 渲染
# =========================
def render_sheet(ws,
                 rows: List[Dict[str, Any]],
                 project_name: str,
                 span: Tuple[Optional[dt.date], Optional[dt.date]],
                 single_day: bool,
                 sheet_title_suffix: str,
                 sheet_kind: str):
    # 针对子表构建列清单
    sheet_specs = build_columns_for_sheet(sheet_kind)
    ncols = len(sheet_specs)

    # 列宽
    for idx, w in enumerate(col_widths_for_specs(ncols), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    if not (span[0] and span[1]) and rows:
        span = detect_date_span_from_rows(rows); single_day = (span[0] and span[1] and span[0]==span[1])
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        f"{project_name}CATATAN ANALISIS OTOMATIS AI DARI DATA KENDARAAN - {sheet_title_suffix}\n"
        f"车联网数据AI自动分析记录 - {sheet_title_suffix}\n"
        f"{('（%s）' % span[0].strftime('%Y.%m.%d')) if (span[0] and span[1] and single_day) else (f'（{span[0].strftime('%Y.%m.%d')}-{span[1].strftime('%Y.%m.%d')}）' if span[0] and span[1] else '（）')}"
    )
    title_cell.font = Font(name=FONT_TITLE_NAME, size=26, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 110

    # 表头
    ws.row_dimensions[2].height = 40
    head_font = Font(name=FONT_BODY_NAME, size=10, bold=True)
    head_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j,(f,zh,_) in enumerate(sheet_specs, start=1):
        cell = ws.cell(row=2, column=j, value=f"{f}\n{zh}")
        cell.font=head_font; cell.alignment=head_align

    # 正文
    body_font = Font(name=FONT_BODY_NAME, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left = Alignment(horizontal="left",   vertical="top",   wrap_text=True)
    thin, thick = Side(style="thin", color="000000"), Side(style="thick", color="000000")
    border_all = Border(left=thin,right=thin,top=thin,bottom=thin)

    # 找出当前表的 SUMMARY 与 PURCH_PRICE 的列号，便于设置样式
    summary_col_idx = next((idx+1 for idx,(en,_,_) in enumerate(sheet_specs) if en=="SUMMARY"), None)
    purch_col_idx   = next((idx+1 for idx,(en,_,_) in enumerate(sheet_specs) if en=="PURCH_PRICE"), None)

    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 40
        for c_idx, (_,_,alts) in enumerate(sheet_specs, start=1):
            v = _first_present(row, alts)
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = body_font
            cell.alignment = top_left if (summary_col_idx is not None and c_idx==summary_col_idx) else center
            cell.border = border_all

    max_row = max(2, 2+len(rows))
    for r in range(1, max_row+1):
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c)
            left   = thick if c==1       else cell.border.left
            right  = thick if c==ncols   else cell.border.right
            top    = thick if r==1       else cell.border.top
            bottom = thick if r==max_row else cell.border.bottom
            cell.border = Border(left=left,right=right,top=top,bottom=bottom)

    # 价格列格式
    if purch_col_idx is not None:
        for r in range(3, max_row+1):
            ws.cell(row=r, column=purch_col_idx).number_format = 'Rp#,##0.00'

def make_excel_multisheet(rows: List[Dict[str, Any]], project_name: str,
                          span: Tuple[Optional[dt.date], Optional[dt.date]], single_day: bool) -> bytes:
    # 分类
    sedang_rows, berat_rows = [], []
    for r in rows:
        cat = classify_alat(r)
        if cat == "sedang":
            sedang_rows.append(r)
        elif cat == "berat":
            berat_rows.append(r)

    # 排名（当日）
    sedang_ranked = sort_and_rank_by_date(sedang_rows)
    berat_ranked  = sort_and_rank_by_date(berat_rows)

    wb = Workbook()
    # 第1张：Alat Sedang
    ws1 = wb.active
    ws1.title = "Alat Sedang"
    render_sheet(ws1, sedang_ranked, project_name, span, single_day,
                 "Alat Sedang (Peringkat Harian)", sheet_kind="sedang")

    # 第2张：Alat Berat
    ws2 = wb.create_sheet(title="Alat Berat")
    render_sheet(ws2, berat_ranked, project_name, span, single_day,
                 "Alat Berat (Peringkat Harian)", sheet_kind="berat")

    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return bio.read()

# =========================
# 可选：本地简易表单
# =========================
@app.get("/", response_class=HTMLResponse)
@app.get("/ui", response_class=HTMLResponse)
def ui(error: Optional[str] = None):
    return """
<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI 报表下载</title></head>
<body style="font-family:Segoe UI,Roboto,Arial,'微软雅黑';max-width:880px;margin:40px auto;padding:0 16px">
<h1>车联网数据 AI 报表下载</h1>
<p>导出的Excel包含两张子表：Alat Sedang（第1张）与 Alat Berat（第2张）；每张表的显示列与顺序按你的要求定制，且为“当日排名”。</p>
<div>
<label>单日：</label><input id="DATE_STR" placeholder="YYYY-MM-DD">
<label style="margin-left:12px">开始：</label><input id="DATE_FROM" placeholder="YYYY-MM-DD">
<label style="margin-left:12px">结束：</label><input id="DATE_TO" placeholder="YYYY-MM-DD">
</div>
<div style="margin-top:8px">
<label>项目名：</label><input id="PROJECT_NAME" placeholder="默认 Kalteng GIJ 中加一园" style="width:60%">
</div>
<button onclick="go()" style="margin-top:10px">下载 Excel</button>
<script>
function enc(s){return encodeURIComponent((s||'').trim());}
function go(){
  const d=document.getElementById('DATE_STR').value;
  const f=document.getElementById('DATE_FROM').value;
  const t=document.getElementById('DATE_TO').value;
  const p=document.getElementById('PROJECT_NAME').value;
  let qs=[];
  if(d){qs.push('DATE_STR='+enc(d));}
  else if(f&&t){qs.push('DATE_FROM='+enc(f));qs.push('DATE_TO='+enc(t));}
  else{alert('请填写 单日 或 起止日期');return;}
  if(p) qs.push('PROJECT_NAME='+enc(p));
  location.href='/download_excel?'+qs.join('&');
}
</script></body></html>
"""

# =========================
# API
# =========================
@app.get("/healthz")
def healthz():
    try:
        conn = connect_db(); conn.close()
        return {"ok": True, "table": DB_TABLE}
    except Exception as e:
        return {"ok": False, "error": str(e), "table": DB_TABLE}

@app.get("/download_excel")
def download_excel(
    request: Request,
    DATE_STR: Optional[str] = Query(default=None, description="单日，例如 2025-10-05"),
    DATE_FROM: Optional[str] = Query(default=None, description="起始日，例如 2025-10-01"),
    DATE_TO: Optional[str]   = Query(default=None, description="结束日，例如 2025-10-07"),
    DATE_RANGE: Optional[str]= Query(default=None, description="范围，如 '2025-10-01到2025-10-07'"),
    PROJECT_NAME: Optional[str] = Query(default=None, description="项目名；缺省为 Kalteng GIJ 中加一园"),
    date: Optional[str] = Query(default=None, alias="date", description="兼容旧参数，等价于 DATE_STR"),
):
    # 参数规范化
    DATE_STR   = DATE_STR.strip() if DATE_STR else None
    DATE_FROM  = DATE_FROM.strip() if DATE_FROM else None
    DATE_TO    = DATE_TO.strip() if DATE_TO else None
    DATE_RANGE = DATE_RANGE.strip() if DATE_RANGE else None
    date = date.strip() if date else None
    if date and not DATE_STR:
        DATE_STR = date

    PROJECT_NAME = normalize_project_name(PROJECT_NAME)
    input_span = span_from_params(DATE_STR, DATE_FROM, DATE_TO, DATE_RANGE)

    # 时间必填
    span = (None, None); single_day = False
    where, params = [], []

    if DATE_STR:
        d = parse_date_like(DATE_STR)
        if not d:
            msg = "DATE_STR 格式不合法，应为 YYYY-MM-DD / YYYY.MM.DD / YYYY/MM/DD"
            return redirect_with_error(msg,
                {"DATE_STR": DATE_STR, "PROJECT_NAME": PROJECT_NAME}, input_span) if want_html(request) \
                else JSONResponse(status_code=400, content={"error": msg})
        where.append("DATE_STR = %s"); params.append(d.strftime("%Y-%m-%d"))
        span, single_day = (d, d), True
    else:
        left = parse_date_like(DATE_FROM) if DATE_FROM else None
        right = parse_date_like(DATE_TO) if DATE_TO else None
        if not (left and right) and DATE_RANGE:
            l2, r2 = parse_date_range(DATE_RANGE)
            left, right = left or l2, right or r2
        if left and right:
            where.append("DATE_STR BETWEEN %s AND %s")
            params += [left.strftime("%Y-%m-%d"), right.strftime("%Y-%m-%d")]
            span, single_day = (left, right), (left == right)
        else:
            msg = "时间必填：传 DATE_STR 或 DATE_FROM+DATE_TO 或 DATE_RANGE"
            return redirect_with_error(msg,
                {"DATE_FROM": DATE_FROM, "DATE_TO": DATE_TO, "DATE_RANGE": DATE_RANGE, "PROJECT_NAME": PROJECT_NAME},
                input_span) if want_html(request) \
                else JSONResponse(status_code=400, content={"error": msg})

    # 项目
    project_for_title = PROJECT_NAME or DEFAULT_PROJECT
    where.append("PROJECT_NAME = %s"); params.append(project_for_title)

    sql = f"SELECT * FROM `{DB_TABLE}` WHERE " + " AND ".join(where) + " ORDER BY DATE_STR ASC, ID ASC"

    try:
        conn = connect_db()
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall() or []
    except Exception as e:
        msg = f"DB error: {e}"
        return redirect_with_error(msg,
            {"DATE_STR": DATE_STR, "DATE_FROM": DATE_FROM, "DATE_TO": DATE_TO, "DATE_RANGE": DATE_RANGE, "PROJECT_NAME": project_for_title},
            input_span) if want_html(request) \
            else JSONResponse(status_code=500, content={"error": msg})
    finally:
        try: conn.close()
        except: pass

    if not rows:
        msg = "没有匹配数据"
        return redirect_with_error(msg,
            {"DATE_STR": DATE_STR, "DATE_FROM": DATE_FROM, "DATE_TO": DATE_TO, "DATE_RANGE": DATE_RANGE, "PROJECT_NAME": project_for_title},
            input_span) if want_html(request) \
            else JSONResponse(status_code=404, content={"error": msg})

    # 生成 Excel（多子表，含“当日排名”，各自定制列）
    excel_bytes = make_excel_multisheet(rows, project_for_title, span, single_day)

    # 文件名
    if span[0] and span[1]:
        date_tag = span[0].strftime("%Y%m%d") if single_day else f"{span[0].strftime('%Y%m%d')}-{span[1].strftime('%Y%m%d')}"
    else:
        date_tag = dt.datetime.now().strftime("%Y%m%d")
    safe_proj = re.sub(r"[\\/:*?\"<>|]+", "_", project_for_title)
    filename = f"{safe_proj}_AI_REPORT_{date_tag}.xlsx"
    ascii_fallback = re.sub(r"[^\x00-\x7F]+", "_", filename)
    utf8_filename = quote(filename)
    headers = {"Content-Disposition": f"attachment; filename={ascii_fallback}; filename*=UTF-8''{utf8_filename}"}

    return StreamingResponse(io.BytesIO(excel_bytes),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)

if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    uvicorn_run("main:app", host="0.0.0.0", port=port, reload=True)
